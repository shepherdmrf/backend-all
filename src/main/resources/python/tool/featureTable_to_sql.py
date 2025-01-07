import os

models=["GPR","KNN","MLR","SVR"]

from openpyxl import load_workbook
import pymysql

for model in models:
    # 加载 Excel 文件
    file = "./src/main/resources/python/MultifacetedModeling/Results/OnTrain/DKNCOR/" + model + ".xlsx"
    wb = load_workbook(file)

    all_data = []

    # 获取工作表数量
    sheet_count = len(wb.worksheets)

    # 遍历所有工作表
    for sheet_index in range(sheet_count):
        data_sheet = wb.worksheets[sheet_index]
        # 遍历工作表中的数据
        tag_name = "result" + str(sheet_index+1)
        for row_index in range(2, data_sheet.max_row+1):
            row_list = []
            for cell in data_sheet[row_index]:
                value = cell.value
                # 根据需要转换数据类型
                if isinstance(value, float):
                    value = round(value, 4)
                elif value is None:
                    value = None
                else:
                    value = str(value)
                row_list.append(value)
            row_list.append(tag_name)
            all_data.append(row_list)

    # 连接数据库
    conn = pymysql.connect(host='localhost', user='root', passwd='123456', db='feature_table')
    cursor = conn.cursor()

    # 删除所有数据并重置 AUTO_INCREMENT 值
    try:
        cursor.execute(" DELETE FROM "+model+"_feature_table")
        cursor.execute(" ALTER TABLE "+model+"_feature_table AUTO_INCREMENT = 1")
        conn.commit()
        print(" The table is emptied and the AUTO_INCREMENT is reset to 1")
    except Exception as e:
        print(f" An error occurred while resetting the table: {e}")

    # 使用参数化查询插入数据
    s = "INSERT INTO `" + model + "_feature_table` (features, Length, Score, RMSE, `CV RMSE`, R2, Time, d_class, tag) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);"

    for data in all_data:
        try:
            cursor.execute(s, data)
        except Exception as e:
            print(f" An error occurred while inserting data: {e}, 数据: {data}")

    conn.commit()

    print(model+" All data in the decision table is inserted successfully! \n")

    cursor.close()
    conn.close()



