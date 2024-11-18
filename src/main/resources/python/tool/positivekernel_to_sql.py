models = ["GPR", "KNN", "MLR", "SVR"]
# 加载 Excel 文件
from openpyxl import load_workbook
import pymysql
def convertPositiveKernelTosql():
    model_lists = {}  # 使用字典来存储模型列表
    for model in models:
        file = f"../MultifacetedModeling/DataOutput/{model}/final_kernels.xlsx"
        wb = load_workbook(file)
        data_sheet = wb.worksheets[0]
        first_column_list = [str(data_sheet.cell(row=i, column=1).value) for i in range(2, data_sheet.max_row + 1)]
        model_lists[model] = first_column_list  # 将模型列表存入字典
    print(model_lists)
    conn = pymysql.connect(host='localhost', user='root', passwd='123456', db='feature_table')
    cursor = conn.cursor()
    for model in models:
        try:
            cursor.execute(f"DELETE FROM {model}_positivekernel")
            cursor.execute(f"ALTER TABLE {model}_positivekernel AUTO_INCREMENT = 1;")
            conn.commit()
            print("表已清空，AUTO_INCREMENT 重置为 1。")
        except Exception as e:
            print(f"重置表时出错: {e}")
        try:
            s = f"INSERT INTO {model}_positivekernel (features) VALUES (%s);"
            for item in model_lists[model]:
                cursor.execute(s,(item))
            print("成功插入")
        except Exception as e:
            print(f"插入表时出错: {e}")
    conn.commit()
    conn.close()
convertPositiveKernelTosql()
