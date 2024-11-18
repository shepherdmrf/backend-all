models = ["GPR", "KNN", "MLR", "SVR"]
# 加载 Excel 文件
from openpyxl import load_workbook
import pymysql

def convertKnowlegeKernelTosql(KnowlegeKernel):
    model_lists = {}  # 使用字典来存储模型列表
    for model in models:
        model_list=[]
        for i in range(len(KnowlegeKernel)):
            file = f"../MultifacetedModeling/Results/OnTrain/KnowledgeKernel/{model}/{i}.xlsx"
            wb = load_workbook(file)
            data_sheet = wb.worksheets[2]
            first_column_list = [float(data_sheet.cell(row=i, column=1).value) for i in range(2, data_sheet.max_row + 1)]
            second_column_list = [float(data_sheet.cell(row=i, column=2).value) for i in range(2, data_sheet.max_row + 1)]
            third_column_list = [float(data_sheet.cell(row=i, column=3).value) for i in range(2, data_sheet.max_row + 1)]
            model_complex=[]
            model_complex.append(first_column_list)
            model_complex.append(second_column_list)
            model_complex.append(third_column_list)
            model_list.append(model_complex)
        model_lists[model] = model_list  # 将模型列表存入字典

    model_knowledgeKernel= {}
    for model in models:
        temp=[]
        for j in range(10):
            min = 100;
            index = 0;
            list=[];
            for i in range(len(model_lists[model])):
                if model_lists[model][i][0][j] < min:
                    index = i
                    min = model_lists[model][i][0][j]
            list.append(KnowlegeKernel[index])
            list.append(model_lists[model][index][0][j])
            list.append(model_lists[model][index][1][j])
            list.append(model_lists[model][index][2][j])
            temp.append(list)
        model_knowledgeKernel[model]=temp
    print(model_knowledgeKernel)
    conn = pymysql.connect(host='localhost', user='root', passwd='123456', db='feature_table')
    cursor = conn.cursor()
    for model in models:
        try:
            cursor.execute("DELETE FROM " + model + "_knowlegekernel")
            cursor.execute(f"ALTER TABLE {model}_knowlegekernel AUTO_INCREMENT = 1;")
            conn.commit()
            print("表已清空，AUTO_INCREMENT 重置为 1。")
        except Exception as e:
            print(f"重置表时出错: {e}")
        try:
            s = f"INSERT INTO {model}_knowlegekernel (feature,`CV RMSE`,RMSE,R2) VALUES (%s,%s,%s,%s);"
            for item in model_knowledgeKernel[model]:
                st ="["+",".join(str(i) for i in item[0])+"]"
                cursor.execute(s,(st,item[1],item[2],item[3]))
            print("成功插入")
        except Exception as e:
            print(f"插入表时出错: {e}")
    conn.commit()
    conn.close()







convertKnowlegeKernelTosql([[35], [34, 35], [36], [37]])